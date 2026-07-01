from __future__ import annotations

from datetime import UTC, datetime, timedelta
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.dialects import postgresql

from app.models.lead import IndiaMartLead
from app.reports.excel_report import (
    DataFetcher,
    ExcelReportGenerator,
    LeadReportRow,
    ReportFetchWindow,
)


class FakeDataFetcher:
    recorded_at: datetime | None = None

    def fetch_leads(self) -> list[LeadReportRow]:
        now = datetime.now(UTC)
        return [
            LeadReportRow(
                lead_id="L-001",
                product_name="Hydraulic Pump",
                product_category="Industrial Pumps",
                quantity="10 Pieces",
                order_value=Decimal("50000"),
                lead_created_date=now,
                lead_updated_date=now,
                lead_extracted_at=now,
                buyer_name="Rahul Sharma",
                business_name="Sharma Industries",
                city="Pune",
                state="Maharashtra",
                years_active=7,
                phone_available=True,
                email_available=True,
                whatsapp_available=True,
                address_available=True,
                business_verified=True,
                requirements_count=14,
                replies_count=3,
                qualification_score=92,
                lead_category="HOT",
                qualification_reasons=["Direct contact channel is available."],
                contacted=True,
                whatsapp_sent=True,
                catalog_shared=False,
                follow_up_done=False,
                negotiation_started=True,
                order_received=False,
                converted=False,
                last_activity=now - timedelta(days=1),
                phone_number="+91 9876543210",
                email="rahul@example.com",
                whatsapp="+91 9876543210",
                business="Sharma Industries",
                address="Pune Industrial Area",
                gst="27ABCDE1234F1Z5",
                gst_available=True,
            ),
            LeadReportRow(
                lead_id="L-002",
                product_name="Safety Gloves",
                product_category="Safety",
                quantity="100 Boxes",
                order_value=Decimal("12000"),
                lead_created_date=now,
                lead_updated_date=now,
                lead_extracted_at=now,
                buyer_name="Anita Rao",
                business_name="Rao Trading",
                city="Mumbai",
                state="Maharashtra",
                years_active=1,
                phone_available=False,
                email_available=True,
                whatsapp_available=False,
                address_available=False,
                business_verified=False,
                requirements_count=2,
                replies_count=0,
                qualification_score=42,
                lead_category="COLD",
                qualification_reasons=["Insufficient qualification signals available."],
                contacted=False,
                whatsapp_sent=False,
                catalog_shared=False,
                follow_up_done=False,
                negotiation_started=False,
                order_received=False,
                converted=True,
                last_activity=now - timedelta(days=9),
            ),
        ]

    def record_excel_extraction_time(self, extracted_at: datetime | None = None) -> None:
        self.recorded_at = extracted_at


def test_excel_report_generator_creates_required_workbook(tmp_path) -> None:
    output_path = tmp_path / "Lead_Intelligence_Report.xlsx"
    generator = ExcelReportGenerator(output_path=output_path, data_fetcher=FakeDataFetcher())

    result = generator.generate_report()

    workbook = load_workbook(result)
    assert result == output_path
    assert workbook.sheetnames == [
        "Executive Summary",
        "Lead Details",
        "Hot Leads",
        "Follow-Up Queue",
        "Product Analytics",
        "Conversion Analytics",
        "AI Recommendations",
    ]
    assert workbook["Executive Summary"]["B6"].value == 2
    assert workbook["Lead Details"]["I6"].value == "HOT"
    assert workbook["Lead Details"]["K5"].value == "Extracted At"
    assert isinstance(workbook["Lead Details"]["K6"].value, datetime)
    assert workbook["Lead Details"]["L6"].value == "+91 9876543210"
    assert workbook["Lead Details"]["M6"].value == "Yes"
    assert workbook["Lead Details"]["N6"].value == "rahul@example.com"
    assert workbook["Lead Details"]["O6"].value == "Yes"
    assert workbook["Lead Details"]["V6"].value == "27ABCDE1234F1Z5"
    assert workbook["Lead Details"]["W6"].value == "Yes"
    assert workbook["Hot Leads"]["A6"].value == "L-001"
    assert workbook["AI Recommendations"]["A12"].value == "L-001"
    assert generator.last_generated_leads
    assert generator.data_fetcher.recorded_at is not None


def test_data_fetcher_uses_stored_indiamart_score_fields() -> None:
    now = datetime.now(UTC)
    lead = IndiaMartLead(
        lead_fingerprint="abc123",
        external_lead_id="IM-001",
        product_name="Hydraulic Pump",
        product_category="Industrial Pumps",
        quantity="5 Pieces",
        order_value="Rs. 25,000",
        buyer_name="Rahul Sharma",
        business_name="Sharma Industries",
        phone_available=True,
        email_available=True,
        whatsapp_available=True,
        business_available=True,
        address_available=True,
        years_active=6,
        requirements_count=10,
        replies_count=4,
        city="Pune",
        state="Maharashtra",
        raw_payload={
            "qualification_score": 55,
            "contacted": True,
            "phone_number": "+91 9876543210",
            "email": "rahul@example.com",
            "whatsapp_number": "+91 9876543210",
            "address": "Pune Industrial Area",
            "gstin": "27ABCDE1234F1Z5",
        },
        lead_score_value=88,
        lead_score_category="HOT",
        lead_score_reasons=["High buying intent", "Direct contact available"],
        lead_score_explanation={},
        extracted_at=now,
        created_at=now,
        updated_at=now,
    )

    row = DataFetcher(database_url="postgresql+psycopg://unused")._to_report_row(lead)

    assert row.qualification_score == 88
    assert row.lead_category == "HOT"
    assert row.lead_extracted_at == now
    assert row.qualification_reasons == ["High buying intent", "Direct contact available"]
    assert row.order_value == Decimal("25000")
    assert row.contacted is True
    assert row.phone_number == "+91 9876543210"
    assert row.email == "rahul@example.com"
    assert row.whatsapp == "+91 9876543210"
    assert row.address == "Pune Industrial Area"
    assert row.gst == "27ABCDE1234F1Z5"
    assert row.gst_available is True


def test_data_fetcher_filters_first_report_to_three_hour_window() -> None:
    now = datetime(2026, 7, 1, 12, 0, tzinfo=UTC)
    window = ReportFetchWindow(
        generated_at=now,
        lookback_started_at=now - timedelta(hours=3),
        previous_extracted_at=None,
    )

    statement = DataFetcher._lead_statement(window)
    compiled = str(statement.compile(dialect=postgresql.dialect()))

    assert "indiamart_leads.extracted_at >= " in compiled
    assert "indiamart_leads.created_at > " not in compiled


def test_data_fetcher_filters_subsequent_reports_by_three_hours_or_previous_run() -> None:
    now = datetime(2026, 7, 1, 12, 0, tzinfo=UTC)
    window = ReportFetchWindow(
        generated_at=now,
        lookback_started_at=now - timedelta(hours=3),
        previous_extracted_at=now - timedelta(hours=8),
    )

    statement = DataFetcher._lead_statement(window)
    compiled = str(statement.compile(dialect=postgresql.dialect()))

    assert "indiamart_leads.extracted_at >= " in compiled
    assert " OR " in compiled
    assert "indiamart_leads.created_at > " in compiled
