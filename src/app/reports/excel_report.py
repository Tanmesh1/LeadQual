from __future__ import annotations

import logging
import re
from collections import Counter
from collections.abc import Iterable
from copy import copy
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import create_engine, or_, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.lead import IndiaMartLead, ReportExtractionState

LOGGER = logging.getLogger(__name__)

DEFAULT_OUTPUT_PATH = Path("Lead_Intelligence_Report.xlsx")
DEFAULT_REPORT_NAME = DEFAULT_OUTPUT_PATH.name
DEFAULT_INCREMENTAL_LOOKBACK = timedelta(hours=3)


@dataclass(frozen=True)
class LeadReportRow:
    lead_id: str
    product_name: str
    product_category: str
    quantity: str
    order_value: Decimal
    lead_created_date: datetime
    lead_updated_date: datetime
    lead_extracted_at: datetime
    buyer_name: str
    business_name: str
    city: str
    state: str
    years_active: int
    phone_available: bool
    email_available: bool
    whatsapp_available: bool
    address_available: bool
    business_verified: bool
    requirements_count: int
    replies_count: int
    qualification_score: int
    lead_category: str
    qualification_reasons: list[str]
    contacted: bool
    whatsapp_sent: bool
    catalog_shared: bool
    follow_up_done: bool
    negotiation_started: bool
    order_received: bool
    converted: bool
    last_activity: datetime
    phone_number: str = ""
    email: str = ""
    whatsapp: str = ""
    business: str = ""
    address: str = ""
    gst: str = ""
    gst_available: bool = False

    @property
    def days_since_last_activity(self) -> int:
        now = datetime.now(UTC)
        last_activity = _ensure_aware(self.last_activity)
        return max(0, (now - last_activity).days)


@dataclass(frozen=True)
class ReportFetchWindow:
    generated_at: datetime
    lookback_started_at: datetime
    previous_extracted_at: datetime | None


class DataFetcher:
    """Fetches and normalizes lead intelligence data from PostgreSQL."""

    def __init__(
        self,
        database_url: str | None = None,
        *,
        connect_timeout: int = 10,
        report_name: str = DEFAULT_REPORT_NAME,
        incremental_lookback: timedelta = DEFAULT_INCREMENTAL_LOOKBACK,
    ) -> None:
        self.database_url = database_url or settings.sync_database_url
        self.connect_timeout = connect_timeout
        self.report_name = report_name
        self.incremental_lookback = incremental_lookback

    def fetch_leads(self) -> list[LeadReportRow]:
        engine = None
        try:
            engine = create_engine(
                self.database_url,
                pool_pre_ping=True,
                connect_args={"connect_timeout": self.connect_timeout},
            )
            with Session(engine) as session:
                window = self._fetch_window(session)
                statement = self._lead_statement(window)
                rows = session.scalars(
                    statement.order_by(IndiaMartLead.extracted_at.desc())
                ).all()
        except Exception:
            LOGGER.exception("Failed to fetch IndiaMART leads for Excel report")
            raise
        finally:
            if engine is not None:
                engine.dispose()

        return [self._to_report_row(row) for row in rows]

    def record_excel_extraction_time(self, extracted_at: datetime | None = None) -> None:
        engine = None
        try:
            engine = create_engine(
                self.database_url,
                pool_pre_ping=True,
                connect_args={"connect_timeout": self.connect_timeout},
            )
            with Session(engine) as session:
                state = session.scalar(
                    select(ReportExtractionState).where(
                        ReportExtractionState.report_name == self.report_name
                    )
                )
                timestamp = _ensure_aware(extracted_at or datetime.now(UTC))
                if state is None:
                    state = ReportExtractionState(
                        report_name=self.report_name,
                        last_extracted_at=timestamp,
                    )
                    session.add(state)
                else:
                    state.last_extracted_at = timestamp
                session.commit()
        except Exception:
            LOGGER.exception("Failed to store Excel report extraction time")
            raise
        finally:
            if engine is not None:
                engine.dispose()

    def _fetch_window(self, session: Session) -> ReportFetchWindow:
        generated_at = datetime.now(UTC)
        previous_extracted_at = session.scalar(
            select(ReportExtractionState.last_extracted_at).where(
                ReportExtractionState.report_name == self.report_name
            )
        )
        return ReportFetchWindow(
            generated_at=generated_at,
            lookback_started_at=generated_at - self.incremental_lookback,
            previous_extracted_at=(
                _ensure_aware(previous_extracted_at) if previous_extracted_at else None
            ),
        )

    @staticmethod
    def _lead_statement(window: ReportFetchWindow):
        statement = select(IndiaMartLead)
        if window.previous_extracted_at is None:
            return statement.where(IndiaMartLead.extracted_at >= window.lookback_started_at)
        return statement.where(
            or_(
                IndiaMartLead.extracted_at >= window.lookback_started_at,
                IndiaMartLead.created_at > window.previous_extracted_at,
            )
        )

    def _to_report_row(self, lead: IndiaMartLead) -> LeadReportRow:
        payload = lead.raw_payload or {}
        score = lead.lead_score_value
        if score is None:
            score = _int_from_payload(payload, "qualification_score", "score")
        if score is None:
            score = self._derive_score(lead)
        score = max(0, min(100, int(score)))

        lead_category = str(lead.lead_score_category or "").upper()
        if lead_category not in {"HOT", "WARM", "COLD"}:
            lead_category = _category(score)

        return LeadReportRow(
            lead_id=str(lead.external_lead_id or lead.id),
            product_name=lead.product_name or "Unknown Product",
            product_category=lead.product_category or "Uncategorized",
            quantity=lead.quantity or "",
            order_value=_parse_money(lead.order_value),
            lead_created_date=_ensure_aware(lead.created_at or lead.extracted_at),
            lead_updated_date=_ensure_aware(lead.updated_at or lead.extracted_at),
            lead_extracted_at=_ensure_aware(lead.extracted_at),
            buyer_name=lead.buyer_name or "",
            business_name=lead.business_name or lead.buyer_name or "Unknown Business",
            city=lead.city or "Unknown",
            state=lead.state or "",
            years_active=lead.years_active or 0,
            phone_available=bool(lead.phone_available),
            email_available=bool(lead.email_available),
            whatsapp_available=bool(lead.whatsapp_available),
            address_available=bool(lead.address_available),
            business_verified=bool(lead.business_available),
            requirements_count=lead.requirements_count or 0,
            replies_count=lead.replies_count or 0,
            qualification_score=score,
            lead_category=lead_category,
            qualification_reasons=self._qualification_reasons(lead, payload),
            contacted=_payload_bool(payload, "contacted"),
            whatsapp_sent=_payload_bool(payload, "whatsapp_sent"),
            catalog_shared=_payload_bool(payload, "catalog_shared"),
            follow_up_done=_payload_bool(payload, "follow_up_done"),
            negotiation_started=_payload_bool(payload, "negotiation_started"),
            order_received=_payload_bool(payload, "order_received"),
            converted=_payload_bool(payload, "converted"),
            last_activity=_payload_datetime(payload, "last_activity_at")
            or _ensure_aware(lead.updated_at or lead.extracted_at),
            phone_number=_payload_text(
                payload,
                "phone_number",
                "phone",
                "mobile_number",
                "mobile",
                "contact_number",
            ),
            email=_payload_text(payload, "email", "email_address", "buyer_email"),
            whatsapp=_payload_text(
                payload,
                "whatsapp_number",
                "whatsapp",
                "whatsapp_contact",
            ),
            business=_payload_text(payload, "business", "business_name", "company_name")
            or lead.business_name
            or "",
            address=_payload_text(
                payload,
                "address",
                "buyer_address",
                "business_address",
                "location_address",
            ),
            gst=_payload_text(payload, "gst", "gstin", "gst_number", "gst_no", "gstin_number"),
            gst_available=_payload_presence(
                payload,
                "gst_available",
                "gst_present",
                "gst",
                "gstin",
                "gst_number",
                "gst_no",
                "gstin_number",
            ),
        )

    def _qualification_reasons(self, lead: IndiaMartLead, payload: dict[str, Any]) -> list[str]:
        return (
            _normalize_reasons(lead.lead_score_reasons)
            or _reasons(lead.lead_score_explanation or {})
            or _reasons(payload)
            or self._derive_reasons(lead)
        )

    @staticmethod
    def _derive_score(lead: IndiaMartLead) -> int:
        score = 0
        score += 12 if lead.phone_available else 0
        score += 10 if lead.email_available else 0
        score += 10 if lead.whatsapp_available else 0
        score += 12 if lead.business_available else 0
        score += 8 if lead.address_available else 0
        score += min(15, (lead.years_active or 0) * 2)
        score += min(18, (lead.requirements_count or 0) * 2)
        score += min(10, (lead.replies_count or 0) * 2)
        score += 5 if _parse_money(lead.order_value) > 0 else 0
        return max(0, min(100, score))

    @staticmethod
    def _derive_reasons(lead: IndiaMartLead) -> list[str]:
        reasons: list[str] = []
        if lead.business_available:
            reasons.append("Business profile is verified or available.")
        if lead.phone_available or lead.whatsapp_available:
            reasons.append("Direct contact channel is available.")
        if (lead.requirements_count or 0) >= 5:
            reasons.append("Buyer has multiple posted requirements.")
        if (lead.years_active or 0) >= 3:
            reasons.append("Business has operating history.")
        if _parse_money(lead.order_value) > 0:
            reasons.append("Order value indicates potential revenue.")
        return reasons or ["Insufficient qualification signals available."]


class ChartBuilder:
    def add_bar_chart(
        self,
        worksheet,
        *,
        title: str,
        data_range: str,
        category_range: str,
        anchor: str,
        y_axis_title: str = "Count",
    ) -> None:
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = y_axis_title
        chart.add_data(Reference(worksheet, range_string=data_range), titles_from_data=True)
        chart.set_categories(Reference(worksheet, range_string=category_range))
        chart.height = 7
        chart.width = 12
        worksheet.add_chart(chart, anchor)

    def add_pie_chart(
        self,
        worksheet,
        *,
        title: str,
        data_range: str,
        category_range: str,
        anchor: str,
    ) -> None:
        chart = PieChart()
        chart.title = title
        chart.add_data(Reference(worksheet, range_string=data_range), titles_from_data=True)
        chart.set_categories(Reference(worksheet, range_string=category_range))
        chart.height = 7
        chart.width = 10
        worksheet.add_chart(chart, anchor)

    def add_line_chart(
        self,
        worksheet,
        *,
        title: str,
        data_range: str,
        category_range: str,
        anchor: str,
        y_axis_title: str = "Converted Leads",
    ) -> None:
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = y_axis_title
        chart.add_data(Reference(worksheet, range_string=data_range), titles_from_data=True)
        chart.set_categories(Reference(worksheet, range_string=category_range))
        chart.height = 7
        chart.width = 12
        worksheet.add_chart(chart, anchor)


class SheetBuilder:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    def __init__(self, chart_builder: ChartBuilder | None = None) -> None:
        self.charts = chart_builder or ChartBuilder()

    def add_title(self, worksheet, title: str, *, end_column: str = "H") -> None:
        worksheet.merge_cells(f"A1:{end_column}1")
        worksheet["A1"] = "Lead Intelligence Report"
        worksheet["A1"].fill = self.title_fill
        worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet.merge_cells(f"A2:{end_column}2")
        worksheet["A2"] = title
        worksheet["A2"].fill = self.subtitle_fill
        worksheet["A2"].font = Font(bold=True, color="1F4E78")
        worksheet["A2"].alignment = Alignment(horizontal="center")
        worksheet["A3"] = "Generated At"
        worksheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def write_table(
        self,
        worksheet,
        *,
        start_row: int,
        headers: list[str],
        rows: list[list[Any]],
        table_name: str | None = None,
    ) -> int:
        for col_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=col_index, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_index, row in enumerate(rows, start=start_row + 1):
            for col_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = self.border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '$#,##0.00'
                elif isinstance(value, datetime):
                    cell.value = _excel_datetime(value)
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

        if table_name and rows:
            end_col = worksheet.cell(row=start_row, column=len(headers)).column_letter
            table_ref = f"A{start_row}:{end_col}{start_row + len(rows)}"
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        return start_row + len(rows)

    def apply_common_formatting(self, worksheet) -> None:
        worksheet.sheet_view.showGridLines = False
        for row in worksheet.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.vertical = "top"
                cell.alignment = alignment
        self.auto_size_columns(worksheet)

    @staticmethod
    def auto_size_columns(worksheet) -> None:
        for column_index in range(1, worksheet.max_column + 1):
            letter = get_column_letter(column_index)
            max_length = max(
                len(str(worksheet.cell(row=row_index, column=column_index).value or ""))
                for row_index in range(1, worksheet.max_row + 1)
            )
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)


class ExcelReportGenerator:
    def __init__(
        self,
        *,
        output_path: str | Path = DEFAULT_OUTPUT_PATH,
        data_fetcher: DataFetcher | None = None,
        sheet_builder: SheetBuilder | None = None,
    ) -> None:
        self.output_path = Path(output_path)
        self.data_fetcher = data_fetcher or DataFetcher()
        self.sheet_builder = sheet_builder or SheetBuilder()
        self.last_generated_leads: list[LeadReportRow] = []

    def generate_report(self) -> Path:
        try:
            generated_at = datetime.now(UTC)
            leads = self.data_fetcher.fetch_leads()
            self.last_generated_leads = leads
            LOGGER.info("Generating Excel lead intelligence report with %s leads", len(leads))
            workbook = Workbook()
            workbook.remove(workbook.active)

            self._build_executive_summary(workbook, leads)
            self._build_lead_details(workbook, leads)
            self._build_hot_leads(workbook, leads)
            self._build_follow_up_queue(workbook, leads)
            self._build_product_analytics(workbook, leads)
            self._build_conversion_analytics(workbook, leads)
            self._build_ai_recommendations(workbook, leads)

            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self.output_path)
            if hasattr(self.data_fetcher, "record_excel_extraction_time"):
                self.data_fetcher.record_excel_extraction_time(generated_at)
            LOGGER.info("Excel lead intelligence report written to %s", self.output_path)
            return self.output_path
        except Exception:
            LOGGER.exception("Failed to generate Excel lead intelligence report")
            raise

    def _build_executive_summary(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Executive Summary")
        self.sheet_builder.add_title(ws, "Executive Summary", end_column="K")

        total = len(leads)
        hot = sum(1 for lead in leads if lead.lead_category == "HOT")
        warm = sum(1 for lead in leads if lead.lead_category == "WARM")
        cold = sum(1 for lead in leads if lead.lead_category == "COLD")
        converted = sum(1 for lead in leads if lead.converted)
        average_score = (
            round(sum(lead.qualification_score for lead in leads) / total, 2) if total else 0
        )
        revenue = sum((lead.order_value for lead in leads), Decimal("0"))

        metrics = [
            ["Total Leads", total],
            ["Hot Leads", hot],
            ["Warm Leads", warm],
            ["Cold Leads", cold],
            ["Average Lead Score", average_score],
            ["Converted Leads", converted],
            ["Conversion Rate", converted / total if total else 0],
            ["Total Potential Revenue", revenue],
        ]
        self.sheet_builder.write_table(ws, start_row=5, headers=["Metric", "Value"], rows=metrics)
        ws["B11"].number_format = "0.00%"
        ws["B12"].number_format = '$#,##0.00'

        self._write_counter_table(
            ws,
            5,
            4,
            "Category",
            Counter(lead.lead_category for lead in leads),
        )
        self._write_counter_table(
            ws,
            5,
            7,
            "Product Category",
            Counter(lead.product_category for lead in leads).most_common(8),
        )
        self._write_counter_table(
            ws,
            18,
            4,
            "City",
            Counter(lead.city for lead in leads).most_common(8),
        )

        self.sheet_builder.charts.add_pie_chart(
            ws,
            title="Hot/Warm/Cold Distribution",
            data_range="'Executive Summary'!$E$5:$E$8",
            category_range="'Executive Summary'!$D$6:$D$8",
            anchor="G14",
        )
        self.sheet_builder.charts.add_bar_chart(
            ws,
            title="Leads By Product Category",
            data_range="'Executive Summary'!$H$5:$H$13",
            category_range="'Executive Summary'!$G$6:$G$13",
            anchor="J5",
        )
        self.sheet_builder.charts.add_bar_chart(
            ws,
            title="Leads By City",
            data_range="'Executive Summary'!$E$18:$E$26",
            category_range="'Executive Summary'!$D$19:$D$26",
            anchor="J20",
        )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_lead_details(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Lead Details")
        self.sheet_builder.add_title(ws, "Lead Details", end_column="W")
        headers = [
            "Lead ID", "Business Name", "Buyer Name", "Product", "Quantity",
            "Order Value", "Years Active", "Score", "Category", "Created Date",
            "Extracted At",
            "Phone Number", "Phone Present", "Email", "Email Present",
            "WhatsApp", "WhatsApp Present", "Business", "Business Present",
            "Address", "Address Present", "GST", "GST Present",
        ]
        rows = [
            [
                lead.lead_id, lead.business_name, lead.buyer_name, lead.product_name,
                lead.quantity, lead.order_value, lead.years_active, lead.qualification_score,
                lead.lead_category, lead.lead_created_date, lead.lead_extracted_at,
                lead.phone_number, _yes_no(lead.phone_available),
                lead.email, _yes_no(lead.email_available),
                lead.whatsapp, _yes_no(lead.whatsapp_available),
                lead.business, _yes_no(lead.business_verified),
                lead.address, _yes_no(lead.address_available),
                lead.gst, _yes_no(lead.gst_available),
            ]
            for lead in leads
        ]
        self.sheet_builder.write_table(
            ws,
            start_row=5,
            headers=headers,
            rows=rows,
            table_name="LeadDetailsTable",
        )
        ws.freeze_panes = "A6"
        end_row = max(6, 5 + len(rows))
        ws.conditional_formatting.add(
            f"I6:I{end_row}",
            CellIsRule(
                operator="equal",
                formula=['"HOT"'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        ws.conditional_formatting.add(
            f"I6:I{end_row}",
            CellIsRule(
                operator="equal",
                formula=['"WARM"'],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        ws.conditional_formatting.add(
            f"I6:I{end_row}",
            CellIsRule(
                operator="equal",
                formula=['"COLD"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_hot_leads(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Hot Leads")
        self.sheet_builder.add_title(ws, "Hot Leads", end_column="F")
        hot_leads = sorted(
            (lead for lead in leads if lead.qualification_score >= 80),
            key=lambda lead: lead.qualification_score,
            reverse=True,
        )
        rows = [
            [
                lead.lead_id,
                lead.business_name,
                lead.product_name,
                lead.quantity,
                lead.qualification_score,
                "\n".join(lead.qualification_reasons),
            ]
            for lead in hot_leads
        ]
        self.sheet_builder.write_table(
            ws,
            start_row=5,
            headers=["Lead ID", "Business Name", "Product", "Quantity", "Score", "Reason"],
            rows=rows,
            table_name="HotLeadsTable" if rows else None,
        )
        end_row = max(6, 5 + len(rows))
        ws.conditional_formatting.add(
            f"E6:E{end_row}",
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["90"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_follow_up_queue(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Follow-Up Queue")
        self.sheet_builder.add_title(ws, "Follow-Up Queue", end_column="E")
        rows = [
            [
                lead.lead_id,
                lead.business_name,
                _activity_label(lead),
                lead.days_since_last_activity,
                _next_action(lead),
            ]
            for lead in sorted(leads, key=lambda item: item.days_since_last_activity, reverse=True)
        ]
        self.sheet_builder.write_table(
            ws,
            start_row=5,
            headers=[
                "Lead ID",
                "Business Name",
                "Last Activity",
                "Days Since Last Activity",
                "Next Recommended Action",
            ],
            rows=rows,
            table_name="FollowUpQueueTable" if rows else None,
        )
        end_row = max(6, 5 + len(rows))
        ws.conditional_formatting.add(
            f"D6:D{end_row}",
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["7"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        ws.conditional_formatting.add(
            f"D6:D{end_row}",
            CellIsRule(
                operator="between",
                formula=["3", "6"],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        ws.conditional_formatting.add(
            f"D6:D{end_row}",
            CellIsRule(
                operator="lessThan",
                formula=["3"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_product_analytics(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Product Analytics")
        self.sheet_builder.add_title(ws, "Product Analytics", end_column="D")
        df = _to_dataframe(leads)
        if df.empty:
            rows: list[list[Any]] = []
        else:
            product_df = (
                df.groupby("product_name", as_index=False)
                .agg(
                    lead_count=("lead_id", "count"),
                    average_score=("qualification_score", "mean"),
                    potential_revenue=("order_value", "sum"),
                )
                .sort_values("lead_count", ascending=False)
            )
            rows = [
                [
                    item.product_name,
                    int(item.lead_count),
                    round(float(item.average_score), 2),
                    Decimal(str(item.potential_revenue)),
                ]
                for item in product_df.itertuples(index=False)
            ]
        self.sheet_builder.write_table(
            ws,
            start_row=5,
            headers=["Product Name", "Lead Count", "Average Score", "Potential Revenue"],
            rows=rows,
            table_name="ProductAnalyticsTable" if rows else None,
        )
        end_row = max(6, 5 + len(rows))
        if rows:
            self.sheet_builder.charts.add_bar_chart(
                ws,
                title="Product Lead Count",
                data_range=f"'Product Analytics'!$B$5:$B${end_row}",
                category_range=f"'Product Analytics'!$A$6:$A${end_row}",
                anchor="F5",
            )
            self.sheet_builder.charts.add_pie_chart(
                ws,
                title="Potential Revenue Share",
                data_range=f"'Product Analytics'!$D$5:$D${end_row}",
                category_range=f"'Product Analytics'!$A$6:$A${end_row}",
                anchor="F20",
            )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_conversion_analytics(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("Conversion Analytics")
        self.sheet_builder.add_title(ws, "Conversion Analytics", end_column="D")
        rows = [
            ["Total Leads", len(leads)],
            ["Qualified Leads", sum(1 for lead in leads if lead.qualification_score >= 50)],
            ["Contacted Leads", sum(1 for lead in leads if lead.contacted)],
            ["Negotiation Leads", sum(1 for lead in leads if lead.negotiation_started)],
            ["Converted Leads", sum(1 for lead in leads if lead.converted)],
        ]
        self.sheet_builder.write_table(ws, start_row=5, headers=["Stage", "Lead Count"], rows=rows)

        trend_rows = self._monthly_conversion_rows(leads)
        self.sheet_builder.write_table(
            ws,
            start_row=13,
            headers=["Month", "Converted Leads"],
            rows=trend_rows,
        )
        self.sheet_builder.charts.add_bar_chart(
            ws,
            title="Conversion Funnel",
            data_range="'Conversion Analytics'!$B$5:$B$10",
            category_range="'Conversion Analytics'!$A$6:$A$10",
            anchor="D5",
        )
        if trend_rows:
            end_row = 13 + len(trend_rows)
            self.sheet_builder.charts.add_line_chart(
                ws,
                title="Monthly Conversion Trend",
                data_range=f"'Conversion Analytics'!$B$13:$B${end_row}",
                category_range=f"'Conversion Analytics'!$A$14:$A${end_row}",
                anchor="D20",
            )
        self.sheet_builder.apply_common_formatting(ws)

    def _build_ai_recommendations(self, workbook: Workbook, leads: list[LeadReportRow]) -> None:
        ws = workbook.create_sheet("AI Recommendations")
        self.sheet_builder.add_title(ws, "AI Recommendations", end_column="E")
        best_category = _best_category(leads)
        highest_quality = max(leads, key=lambda lead: lead.qualification_score, default=None)
        biggest_revenue = max(leads, key=lambda lead: lead.order_value, default=None)
        summary_rows = [
            ["Best Performing Product Category", best_category],
            ["Highest Quality Lead", highest_quality.business_name if highest_quality else "N/A"],
            [
                "Biggest Revenue Opportunity",
                biggest_revenue.business_name if biggest_revenue else "N/A",
            ],
        ]
        self.sheet_builder.write_table(
            ws,
            start_row=5,
            headers=["Executive Insight", "Value"],
            rows=summary_rows,
        )

        hot_leads = sorted(
            (lead for lead in leads if lead.lead_category == "HOT"),
            key=lambda lead: lead.qualification_score,
            reverse=True,
        )
        rows = [
            [
                lead.lead_id,
                lead.business_name,
                lead.qualification_score,
                "\n".join(lead.qualification_reasons),
                "\n".join(_recommended_actions(lead)),
            ]
            for lead in hot_leads
        ]
        self.sheet_builder.write_table(
            ws,
            start_row=11,
            headers=[
                "Lead ID",
                "Business Name",
                "Lead Score",
                "Qualification Reasons",
                "Recommended Actions",
            ],
            rows=rows,
            table_name="AIRecommendationsTable" if rows else None,
        )
        self.sheet_builder.apply_common_formatting(ws)

    @staticmethod
    def _monthly_conversion_rows(leads: list[LeadReportRow]) -> list[list[Any]]:
        converted_leads = [lead for lead in leads if lead.converted]
        counter = Counter(lead.lead_updated_date.strftime("%Y-%m") for lead in converted_leads)
        return [[month, count] for month, count in sorted(counter.items())]

    @staticmethod
    def _write_counter_table(
        worksheet,
        start_row: int,
        start_col: int,
        title: str,
        values: Any,
    ) -> None:
        items = values.items() if hasattr(values, "items") else values
        worksheet.cell(row=start_row, column=start_col, value=title).fill = SheetBuilder.header_fill
        worksheet.cell(
            row=start_row,
            column=start_col + 1,
            value="Lead Count",
        ).fill = SheetBuilder.header_fill
        worksheet.cell(row=start_row, column=start_col).font = SheetBuilder.header_font
        worksheet.cell(row=start_row, column=start_col + 1).font = SheetBuilder.header_font
        for row_index, (name, count) in enumerate(items, start=start_row + 1):
            worksheet.cell(row=row_index, column=start_col, value=name)
            worksheet.cell(row=row_index, column=start_col + 1, value=count)


def generate_report(output_path: str | Path = DEFAULT_OUTPUT_PATH) -> Path:
    return ExcelReportGenerator(output_path=output_path).generate_report()


def _to_dataframe(leads: list[LeadReportRow]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "lead_id": lead.lead_id,
                "product_name": lead.product_name,
                "product_category": lead.product_category,
                "order_value": float(lead.order_value),
                "qualification_score": lead.qualification_score,
                "converted": lead.converted,
            }
            for lead in leads
        ]
    )


def _parse_money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return Decimal("0")
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return Decimal("0")


def _category(score: int) -> str:
    if score >= 80:
        return "HOT"
    if score >= 50:
        return "WARM"
    return "COLD"


def _reasons(payload: dict[str, Any]) -> list[str]:
    reasons = payload.get("qualification_reasons") or payload.get("reasons")
    normalized = _normalize_reasons(reasons)
    if normalized:
        return normalized
    explanation = payload.get("explanation")
    if isinstance(explanation, dict):
        return _normalize_reasons(
            explanation.get("why_score_is_high") or explanation.get("reasons")
        )
    return []


def _normalize_reasons(reasons: Any) -> list[str]:
    if isinstance(reasons, list):
        return [str(reason) for reason in reasons if str(reason).strip()]
    if isinstance(reasons, tuple | set):
        return [str(reason) for reason in reasons if str(reason).strip()]
    if isinstance(reasons, Iterable) and not isinstance(reasons, str | bytes | dict):
        return [str(reason) for reason in reasons if str(reason).strip()]
    if isinstance(reasons, str) and reasons.strip():
        return [reasons.strip()]
    return []


def _payload_bool(payload: dict[str, Any], key: str) -> bool:
    activity = payload.get("activity")
    if isinstance(activity, dict) and key in activity:
        return bool(activity[key])
    return bool(payload.get(key, False))


def _payload_text(payload: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = _find_payload_value(payload, key)
        if value in (None, "", [], {}):
            continue
        if isinstance(value, bool):
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _payload_presence(payload: dict[str, Any], *keys: str) -> bool:
    for key in keys:
        value = _find_payload_value(payload, key)
        if isinstance(value, bool):
            return value
        if value in (None, "", [], {}):
            continue
        text = str(value).strip().lower()
        unavailable_tokens = ("not available", "unavailable", "hidden", "no ", "false")
        if text and not any(token in text for token in unavailable_tokens):
            return True
    return False


def _find_payload_value(payload: dict[str, Any], key: str) -> Any:
    if key in payload:
        return payload[key]

    for nested_key in ("contact", "contacts", "buyer", "business", "company", "raw_payload"):
        nested = payload.get(nested_key)
        if isinstance(nested, dict) and key in nested:
            return nested[key]

    return None


def _yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def _int_from_payload(payload: dict[str, Any], *keys: str) -> int | None:
    for key in keys:
        if key not in payload:
            continue
        try:
            return max(0, min(100, int(Decimal(str(payload[key])))))
        except (InvalidOperation, ValueError, TypeError):
            continue
    return None


def _payload_datetime(payload: dict[str, Any], key: str) -> datetime | None:
    value = payload.get(key) or (payload.get("activity") or {}).get(key)
    if not value:
        return None
    if isinstance(value, datetime):
        return _ensure_aware(value)
    try:
        return _ensure_aware(datetime.fromisoformat(str(value).replace("Z", "+00:00")))
    except ValueError:
        return None


def _ensure_aware(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value


def _excel_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone().replace(tzinfo=None)


def _activity_label(lead: LeadReportRow) -> str:
    if lead.converted:
        return "Converted"
    if lead.order_received:
        return "Order Received"
    if lead.negotiation_started:
        return "Negotiation Started"
    if lead.follow_up_done:
        return "Follow-Up Done"
    if lead.catalog_shared:
        return "Catalog Shared"
    if lead.whatsapp_sent:
        return "WhatsApp Sent"
    if lead.contacted:
        return "Contacted"
    return "No Recorded Activity"


def _next_action(lead: LeadReportRow) -> str:
    if lead.converted:
        return "Maintain account relationship"
    if lead.days_since_last_activity >= 7:
        return "Requires Immediate Follow-Up"
    if lead.days_since_last_activity >= 3:
        return "Follow-Up Soon"
    return "Recently Contacted"


def _recommended_actions(lead: LeadReportRow) -> list[str]:
    actions = ["Call within 30 minutes", "Share wholesale catalog", "Offer MOQ pricing"]
    if not lead.whatsapp_sent and lead.whatsapp_available:
        actions.append("Send WhatsApp follow-up")
    if not lead.catalog_shared:
        actions.append("Send product catalog")
    if lead.order_value > 0:
        actions.append("Prepare value-based quote")
    return actions


def _best_category(leads: list[LeadReportRow]) -> str:
    if not leads:
        return "N/A"
    df = _to_dataframe(leads)
    if df.empty:
        return "N/A"
    grouped = (
        df.groupby("product_category", as_index=False)
        .agg(average_score=("qualification_score", "mean"), lead_count=("lead_id", "count"))
        .sort_values(["average_score", "lead_count"], ascending=False)
    )
    return str(grouped.iloc[0]["product_category"]) if not grouped.empty else "N/A"
